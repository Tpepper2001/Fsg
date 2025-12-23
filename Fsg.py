import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class FinancialStatementGenerator:
    """Generate financial statements from transaction data"""
    
    def __init__(self, transactions_df):
        """
        Initialize with transaction data
        
        Expected columns: date, account, category, amount, type (debit/credit)
        """
        self.df = transactions_df.copy()
        self.df['date'] = pd.to_datetime(self.df['date'])
        self.df['period'] = self.df['date'].dt.to_period('M')
        
    def generate_income_statement(self, period=None):
        """Generate Income Statement for specified period"""
        if period:
            data = self.df[self.df['period'] == period]
        else:
            data = self.df
            
        # Revenue
        revenue = data[data['category'] == 'Revenue']['amount'].sum()
        
        # Cost of Goods Sold
        cogs = data[data['category'] == 'COGS']['amount'].sum()
        gross_profit = revenue - cogs
        
        # Operating Expenses
        operating_exp = data[data['category'].isin([
            'Sales & Marketing', 'General & Administrative', 'R&D'
        ])]['amount'].sum()
        
        operating_income = gross_profit - operating_exp
        
        # Other Income/Expenses
        interest_expense = data[data['category'] == 'Interest Expense']['amount'].sum()
        other_income = data[data['category'] == 'Other Income']['amount'].sum()
        
        pretax_income = operating_income - interest_expense + other_income
        tax_expense = pretax_income * 0.25  # 25% tax rate assumption
        net_income = pretax_income - tax_expense
        
        return pd.DataFrame({
            'Line Item': [
                'Revenue',
                'Cost of Goods Sold',
                'Gross Profit',
                'Operating Expenses',
                'Operating Income',
                'Interest Expense',
                'Other Income',
                'Pre-tax Income',
                'Tax Expense',
                'Net Income'
            ],
            'Amount': [
                revenue,
                -cogs,
                gross_profit,
                -operating_exp,
                operating_income,
                -interest_expense,
                other_income,
                pretax_income,
                -tax_expense,
                net_income
            ]
        })
    
    def generate_balance_sheet(self, as_of_date=None):
        """Generate Balance Sheet as of specified date"""
        if as_of_date:
            data = self.df[self.df['date'] <= as_of_date]
        else:
            data = self.df
            
        # Assets
        cash = data[data['account'] == 'Cash']['amount'].sum()
        ar = data[data['account'] == 'Accounts Receivable']['amount'].sum()
        inventory = data[data['account'] == 'Inventory']['amount'].sum()
        current_assets = cash + ar + inventory
        
        ppe = data[data['account'] == 'PP&E']['amount'].sum()
        total_assets = current_assets + ppe
        
        # Liabilities
        ap = data[data['account'] == 'Accounts Payable']['amount'].sum()
        short_term_debt = data[data['account'] == 'Short-term Debt']['amount'].sum()
        current_liabilities = ap + short_term_debt
        
        long_term_debt = data[data['account'] == 'Long-term Debt']['amount'].sum()
        total_liabilities = current_liabilities + long_term_debt
        
        # Equity
        common_stock = data[data['account'] == 'Common Stock']['amount'].sum()
        retained_earnings = data[data['account'] == 'Retained Earnings']['amount'].sum()
        total_equity = common_stock + retained_earnings
        
        return pd.DataFrame({
            'Line Item': [
                'ASSETS',
                'Current Assets:',
                '  Cash',
                '  Accounts Receivable',
                '  Inventory',
                'Total Current Assets',
                'Fixed Assets:',
                '  PP&E',
                'TOTAL ASSETS',
                '',
                'LIABILITIES & EQUITY',
                'Current Liabilities:',
                '  Accounts Payable',
                '  Short-term Debt',
                'Total Current Liabilities',
                'Long-term Debt',
                'Total Liabilities',
                '',
                'Equity:',
                '  Common Stock',
                '  Retained Earnings',
                'Total Equity',
                '',
                'TOTAL LIABILITIES & EQUITY'
            ],
            'Amount': [
                None,
                None,
                cash,
                ar,
                inventory,
                current_assets,
                None,
                ppe,
                total_assets,
                None,
                None,
                None,
                ap,
                short_term_debt,
                current_liabilities,
                long_term_debt,
                total_liabilities,
                None,
                None,
                common_stock,
                retained_earnings,
                total_equity,
                None,
                total_assets
            ]
        })
    
    def generate_variance_analysis(self, current_period, prior_period):
        """Generate variance analysis between two periods"""
        current = self.generate_income_statement(current_period)
        prior = self.generate_income_statement(prior_period)
        
        variance_df = current.copy()
        variance_df['Prior Period'] = prior['Amount']
        variance_df['Current Period'] = current['Amount']
        variance_df['Variance $'] = variance_df['Current Period'] - variance_df['Prior Period']
        variance_df['Variance %'] = (variance_df['Variance $'] / variance_df['Prior Period'].replace(0, np.nan) * 100).round(1)
        
        return variance_df[['Line Item', 'Prior Period', 'Current Period', 'Variance $', 'Variance %']]
    
    def export_to_excel(self, filename='financial_statements.xlsx'):
        """Export all statements to formatted Excel file"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get current and prior period
        periods = sorted(self.df['period'].unique())
        current_period = periods[-1]
        prior_period = periods[-2] if len(periods) > 1 else periods[-1]
        
        # Income Statement
        ws_is = wb.create_sheet('Income Statement')
        is_df = self.generate_income_statement(current_period)
        self._format_sheet(ws_is, is_df, f'Income Statement - {current_period}')
        
        # Balance Sheet
        ws_bs = wb.create_sheet('Balance Sheet')
        bs_df = self.generate_balance_sheet()
        self._format_sheet(ws_bs, bs_df, f'Balance Sheet - {current_period}')
        
        # Variance Analysis
        ws_var = wb.create_sheet('Variance Analysis')
        var_df = self.generate_variance_analysis(current_period, prior_period)
        self._format_sheet(ws_var, var_df, f'Variance Analysis: {prior_period} vs {current_period}')
        
        wb.save(filename)
        print(f"Financial statements exported to {filename}")
        
    def _format_sheet(self, ws, df, title):
        """Apply formatting to Excel sheet"""
        # Title
        ws.append([title])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        
        # Add dataframe
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format numbers
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def generate_sample_data():
    """Generate sample transaction data for demonstration"""
    np.random.seed(42)
    dates = pd.date_range(start='2024-01-01', end='2024-12-31', freq='D')
    
    transactions = []
    
    # Generate revenue transactions
    for date in dates[::7]:  # Weekly revenue
        transactions.append({
            'date': date,
            'account': 'Cash',
            'category': 'Revenue',
            'amount': np.random.uniform(50000, 100000),
            'type': 'credit'
        })
    
    # Generate COGS
    for date in dates[::7]:
        transactions.append({
            'date': date,
            'account': 'Inventory',
            'category': 'COGS',
            'amount': np.random.uniform(20000, 40000),
            'type': 'debit'
        })
    
    # Operating expenses
    for date in dates[::30]:  # Monthly
        transactions.extend([
            {'date': date, 'account': 'Cash', 'category': 'Sales & Marketing', 
             'amount': np.random.uniform(15000, 25000), 'type': 'debit'},
            {'date': date, 'account': 'Cash', 'category': 'General & Administrative',
             'amount': np.random.uniform(10000, 20000), 'type': 'debit'},
            {'date': date, 'account': 'Cash', 'category': 'R&D',
             'amount': np.random.uniform(8000, 15000), 'type': 'debit'}
        ])
    
    # Balance sheet accounts
    transactions.extend([
        {'date': dates[0], 'account': 'Cash', 'category': 'Asset', 'amount': 500000, 'type': 'debit'},
        {'date': dates[0], 'account': 'Accounts Receivable', 'category': 'Asset', 'amount': 150000, 'type': 'debit'},
        {'date': dates[0], 'account': 'Inventory', 'category': 'Asset', 'amount': 100000, 'type': 'debit'},
        {'date': dates[0], 'account': 'PP&E', 'category': 'Asset', 'amount': 300000, 'type': 'debit'},
        {'date': dates[0], 'account': 'Accounts Payable', 'category': 'Liability', 'amount': 80000, 'type': 'credit'},
        {'date': dates[0], 'account': 'Long-term Debt', 'category': 'Liability', 'amount': 200000, 'type': 'credit'},
        {'date': dates[0], 'account': 'Common Stock', 'category': 'Equity', 'amount': 500000, 'type': 'credit'},
        {'date': dates[0], 'account': 'Retained Earnings', 'category': 'Equity', 'amount': 270000, 'type': 'credit'}
    ])
    
    return pd.DataFrame(transactions)


# Main execution
if __name__ == "__main__":
    # Generate sample data
    print("Generating sample transaction data...")
    transactions = generate_sample_data()
    
    # Save sample data
    transactions.to_csv('sample_transactions.csv', index=False)
    print("Sample data saved to sample_transactions.csv")
    
    # Generate financial statements
    print("\nGenerating financial statements...")
    generator = FinancialStatementGenerator(transactions)
    
    # Display Income Statement
    print("\n=== INCOME STATEMENT ===")
    print(generator.generate_income_statement().to_string(index=False))
    
    # Export to Excel
    generator.export_to_excel('financial_statements.xlsx')
    print("\n✓ Complete! Check financial_statements.xlsx for full reports")
